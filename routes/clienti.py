import os
import json
import openpyxl
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, jsonify
from flask_login import login_required
from werkzeug.utils import secure_filename
import clients as client_loader
from core.pdf_extractor import leggi_pdf

clienti = Blueprint("clienti", __name__, url_prefix="/clienti")


def _get_excel_dir():
    return os.path.join(current_app.config["UPLOAD_FOLDER"], "excel_clienti")


@clienti.route("/")
@login_required
def elenco():
    plugins = client_loader.get_all_plugins()
    return render_template("clienti/elenco.html", plugins=plugins)


@clienti.route("/<id_cliente>")
@login_required
def dettaglio(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
    ddt_importati = []
    if os.path.exists(excel_path):
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            ddt_importati.append(row)
        wb.close()

    return render_template(
        "clienti/dettaglio.html",
        plugin=plugin,
        ddt_importati=ddt_importati,
        excel_path=excel_path,
    )


@clienti.route("/<id_cliente>/upload", methods=["GET", "POST"])
@login_required
def upload_ddt(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    if request.method == "GET":
        return render_template("clienti/upload.html", plugin=plugin)

    files = request.files.getlist("pdf_files")
    if not files or all(f.filename == "" for f in files):
        flash("Nessun file selezionato.", "error")
        return redirect(request.url)

    risultati = []
    errori = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            errori.append(f"{f.filename}: non è un PDF")
            continue

        filename = secure_filename(f"ddt_{id_cliente}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{f.filename}")
        upload_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "clienti", id_cliente)
        os.makedirs(upload_dir, exist_ok=True)
        path_pdf = os.path.join(upload_dir, filename)
        f.save(path_pdf)

        try:
            testo = leggi_pdf(path_pdf)
            dati = plugin.parse_ddt(testo)
            risultati.append({"file": f.filename, "dati": dati, "successo": True})
        except Exception as e:
            errori.append(f"{f.filename}: {str(e)}")
            risultati.append({"file": f.filename, "dati": None, "successo": False, "errore": str(e)})

    if risultati:
        excel_dir = _get_excel_dir()
        os.makedirs(excel_dir, exist_ok=True)
        excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
        dati_validi = [r["dati"] for r in risultati if r["successo"]]
        if dati_validi:
            try:
                plugin.genera_excel(dati_validi, excel_path)
            except Exception as e:
                flash(f"Errore nella generazione Excel: {str(e)}", "error")

    if errori:
        for e in errori:
            flash(e, "error")

    if any(r["successo"] for r in risultati):
        flash(f"{sum(1 for r in risultati if r['successo'])} DDT elaborati con successo.", "success")

    return render_template("clienti/risultati.html", plugin=plugin, risultati=risultati, errori=errori)


@clienti.route("/<id_cliente>/anteprima", methods=["POST"])
@login_required
def anteprima_parser(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        return jsonify({"error": "Cliente non trovato"}), 404

    f = request.files.get("pdf_file")
    if not f or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "File PDF richiesto"}), 400

    upload_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "clienti", "_anteprima")
    os.makedirs(upload_dir, exist_ok=True)
    path_pdf = os.path.join(upload_dir, secure_filename(f.filename))
    f.save(path_pdf)

    try:
        testo = leggi_pdf(path_pdf)
        dati = plugin.parse_ddt(testo)
        return jsonify({"successo": True, "dati": dati, "testo_estratto": testo[:2000]})
    except Exception as e:
        return jsonify({"successo": False, "errore": str(e)}), 400
    finally:
        if os.path.exists(path_pdf):
            os.remove(path_pdf)


@clienti.route("/<id_cliente>/excel")
@login_required
def scarica_excel(id_cliente):
    import shutil
    from pathlib import Path

    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DDT"
        ws.cell(row=1, column=1, value="Nessun DDT importato. Carica i PDF dalla pagina cliente.")
        wb.save(excel_path)

    from flask import send_file
    download_path = os.path.join(excel_dir, f"DDT_{plugin.nome.replace(' ', '_')}_{datetime.now().strftime('%Y%m')}.xlsx")
    shutil.copy2(excel_path, download_path)
    return send_file(download_path, as_attachment=True, download_name=os.path.basename(download_path))