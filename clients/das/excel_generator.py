import os
from datetime import datetime
from openpyxl import load_workbook
from core.excel_writer import apri_o_crea_excel, stile_intestazione, stile_cella, CENTER_ALIGN


class DasExcelWriter:
    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        wb, esistente = apri_o_crea_excel(excel_path)
        if not esistente:
            ws = wb.active
            ws.title = "DDT DAS"
            headers = ["Data", "DDT", "Pallet ID", "Cliente", "Indirizzo", "Peso (Kg)", "Note"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=h)
            stile_intestazione(ws, 1, len(headers))
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 30

        ws = wb["DDT DAS"] if "DDT DAS" in wb.sheetnames else wb.active
        for ddt in ddt_data_list:
            prox_riga = ws.max_row + 1
            stile_cella(ws, prox_riga, 1, ddt.get("data", ""), align=CENTER_ALIGN)
            stile_cella(ws, prox_riga, 2, ddt.get("ddt", ""), align=CENTER_ALIGN)
            pallet_ids = ", ".join(ddt.get("extra", {}).get("pallet_ids", []))
            stile_cella(ws, prox_riga, 3, pallet_ids, align=CENTER_ALIGN)
            stile_cella(ws, prox_riga, 4, ddt.get("cliente", ""))
            stile_cella(ws, prox_riga, 5, ddt.get("extra", {}).get("localita", ""))
            stile_cella(ws, prox_riga, 6, ddt.get("totale_peso", 0), align=CENTER_ALIGN)
            note = ", ".join([a.get("codice", "") for a in ddt.get("articoli", [])])
            stile_cella(ws, prox_riga, 7, note)

        wb.save(excel_path)
        return excel_path