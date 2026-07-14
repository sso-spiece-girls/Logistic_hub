import os
from datetime import datetime
import openpyxl
from core.excel_writer import apri_o_crea_excel, stile_intestazione, stile_cella, CENTER_ALIGN
from openpyxl.utils import get_column_letter


class EllegroupExcelWriter:
    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        wb, esistente = apri_o_crea_excel(excel_path)
        if not esistente:
            ws = wb.active
            ws.title = "DDT ELLE GROUP"
            headers = ["Data", "DDT", "Agente", "Cliente", "Prov.", "Importo", "Note"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=h)
            stile_intestazione(ws, 1, len(headers))
            for col, w in zip("ABCDEFG", [14, 22, 20, 28, 8, 14, 30]):
                ws.column_dimensions[col].width = w

        ws = wb["DDT ELLE GROUP"] if "DDT ELLE GROUP" in wb.sheetnames else wb.active
        for ddt in ddt_data_list:
            prox_riga = ws.max_row + 1
            stile_cella(ws, prox_riga, 1, ddt.get("data", ""), align=CENTER_ALIGN)
            stile_cella(ws, prox_riga, 2, ddt.get("ddt", ""), align=CENTER_ALIGN)
            stile_cella(ws, prox_riga, 3, ddt.get("extra", {}).get("agente", ""))
            stile_cella(ws, prox_riga, 4, ddt.get("cliente", ""))
            stile_cella(ws, prox_riga, 5, ddt.get("extra", {}).get("provincia", ""), align=CENTER_ALIGN)
            stile_cella(ws, prox_riga, 6, ddt.get("extra", {}).get("importo", 0), align=CENTER_ALIGN)
            note = "; ".join([f"{a.get('codice', '')} {a.get('descrizione', '')}" for a in ddt.get("articoli", [])])
            stile_cella(ws, prox_riga, 7, note)

        wb.save(excel_path)
        return excel_path