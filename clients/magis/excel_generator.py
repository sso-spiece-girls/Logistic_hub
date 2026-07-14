import os
from datetime import datetime
from openpyxl import load_workbook
from core.excel_writer import apri_o_crea_excel, stile_intestazione, stile_cella, crea_backup, CENTER_ALIGN


class MagisExcelWriter:
    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        wb, esistente = apri_o_crea_excel(excel_path)
        if not esistente:
            self._crea_struttura(wb)

        for ddt in ddt_data_list:
            self._scrivi_ddt(wb, ddt)

        wb.save(excel_path)
        return excel_path

    def _crea_struttura(self, wb):
        if "BIG BAG" not in [s.title for s in wb.sheetnames]:
            ws = wb.active
            ws.title = "BIG BAG"
            self._inizializza_foglio(ws)
            ws = wb.create_sheet("SMALL BAG")
            self._inizializza_foglio(ws)
            ws = wb.create_sheet("GIACENZE")
            self._inizializza_giacenze(ws)

    def _inizializza_foglio(self, ws):
        headers = ["Data", "DDT", "Cliente", "PLT OUT", "Note"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        stile_intestazione(ws, 1, len(headers))
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 30

    def _inizializza_giacenze(self, ws):
        headers = ["Deposito", "Articolo", "Qta Iniziale", "Uscite", "Qta Finale"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        stile_intestazione(ws, 1, len(headers))

    def _scrivi_ddt(self, wb, ddt):
        for sheet_name in ["BIG BAG", "SMALL BAG"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                prox_riga = ws.max_row + 1
                stile_cella(ws, prox_riga, 1, ddt.get("data", ""), align=CENTER_ALIGN)
                stile_cella(ws, prox_riga, 2, ddt.get("ddt", ""), align=CENTER_ALIGN)
                stile_cella(ws, prox_riga, 3, ddt.get("cliente", ""))
                stile_cella(ws, prox_riga, 4, ddt.get("totale_pallet", 0), align=CENTER_ALIGN)
                note = ", ".join([a.get("codice", "") for a in ddt.get("articoli", [])])
                stile_cella(ws, prox_riga, 5, note)


