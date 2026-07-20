import os
import json
import openpyxl
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, jsonify
from flask_login import login_required
from werkzeug.utils import secure_filename
import clients as client_loader
from core.pdf_extractor import leggi_pdf
from core.auth_decorators import staff_required

clienti = Blueprint("clienti", __name__, url_prefix="/clienti")


def _get_excel_dir():
    return os.path.join(current_app.config["UPLOAD_FOLDER"], "excel_clienti")


@clienti.route("/")
@login_required
def elenco():
    plugins = client_loader.get_all_plugins()
    return render_template("clienti/elenco.html", plugins=plugins)


@clienti.route("/<id_cliente>")
@login_required
def dettaglio(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
    ddt_importati = []
    header_row = []
    if os.path.exists(excel_path):
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        header_row = [c.value for c in ws[1]] if ws.max_row > 0 else []
        for row in ws.iter_rows(min_row=2, values_only=True):
            ddt_importati.append(row)
        wb.close()

    return render_template(
        "clienti/dettaglio.html",
        plugin=plugin,
        ddt_importati=ddt_importati,
        header_row=header_row,
        excel_path=excel_path,
    )


@clienti.route("/<id_cliente>/upload", methods=["GET", "POST"])
@login_required
def upload_ddt(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    if request.method == "GET":
        return render_template("clienti/upload.html", plugin=plugin)

    files = request.files.getlist("pdf_files")
    if not files or all(f.filename == "" for f in files):
        flash("Nessun file selezionato.", "error")
        return redirect(request.url)

    risultati = []
    errori = []
    for f in files:
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            errori.append(f"{f.filename}: non è un PDF")
            continue

        filename = secure_filename(f"ddt_{id_cliente}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{f.filename}")
        upload_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "clienti", id_cliente)
        os.makedirs(upload_dir, exist_ok=True)
        path_pdf = os.path.join(upload_dir, filename)
        f.save(path_pdf)

        try:
            testo = leggi_pdf(path_pdf)
            dati = plugin.parse_ddt(testo)
            risultati.append({"file": f.filename, "dati": dati, "successo": True})
        except Exception as e:
            errori.append(f"{f.filename}: {str(e)}")
            risultati.append({"file": f.filename, "dati": None, "successo": False, "errore": str(e)})

    if risultati:
        excel_dir = _get_excel_dir()
        os.makedirs(excel_dir, exist_ok=True)
        excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
        dati_validi = [r["dati"] for r in risultati if r["successo"]]
        if dati_validi:
            try:
                plugin.genera_excel(dati_validi, excel_path)
            except Exception as e:
                flash(f"Errore nella generazione Excel: {str(e)}", "error")

    if errori:
        for e in errori:
            flash(e, "error")

    if any(r["successo"] for r in risultati):
        flash(f"{sum(1 for r in risultati if r['successo'])} DDT elaborati con successo.", "success")

    return render_template("clienti/risultati.html", plugin=plugin, risultati=risultati, errori=errori)


@clienti.route("/<id_cliente>/anteprima", methods=["POST"])
@login_required
def anteprima_parser(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        return jsonify({"error": "Cliente non trovato"}), 404

    f = request.files.get("pdf_file")
    if not f or not f.filename or not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "File PDF richiesto"}), 400

    upload_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "clienti", "_anteprima")
    os.makedirs(upload_dir, exist_ok=True)
    path_pdf = os.path.join(upload_dir, secure_filename(f.filename))
    f.save(path_pdf)

    try:
        testo = leggi_pdf(path_pdf)
        dati = plugin.parse_ddt(testo)
        return jsonify({"successo": True, "dati": dati, "testo_estratto": testo[:2000]})
    except Exception as e:
        return jsonify({"successo": False, "errore": str(e)}), 400
    finally:
        if os.path.exists(path_pdf):
            os.remove(path_pdf)


@clienti.route("/<id_cliente>/excel")
@login_required
def scarica_excel(id_cliente):
    import shutil
    from pathlib import Path

    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DDT"
        ws.cell(row=1, column=1, value="Nessun DDT importato. Carica i PDF dalla pagina cliente.")
        wb.save(excel_path)

    from flask import send_file
    download_path = os.path.join(excel_dir, f"DDT_{plugin.nome.replace(' ', '_')}_{datetime.now().strftime('%Y%m')}.xlsx")
    shutil.copy2(excel_path, download_path)
    return send_file(download_path, as_attachment=True, download_name=os.path.basename(download_path))


def _modifica_excel(excel_path, row_index, nuovi_valori):
    """Modifica una riga nell'Excel (row_index 0-based, riga 2+ del foglio)."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    excel_row = row_index + 2
    for col_idx, val in enumerate(nuovi_valori, start=1):
        ws.cell(row=excel_row, column=col_idx, value=val)
    wb.save(excel_path)
    wb.close()


def _elimina_excel(excel_path, row_index):
    """Elimina una riga dall'Excel (row_index 0-based, riga 2+ del foglio)."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    excel_row = row_index + 2
    ws.delete_rows(excel_row)
    wb.save(excel_path)
    wb.close()


def _aggiungi_excel(excel_path, nuovi_valori):
    """Aggiunge una riga in coda all'Excel."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append(nuovi_valori)
    wb.save(excel_path)
    wb.close()


@clienti.route("/<id_cliente>/modifica-riga", methods=["POST"])
@login_required
def modifica_riga(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        return jsonify({"error": "Cliente non trovato"}), 404

    data = request.get_json(silent=True) or {}
    row_index = data.get("row_index")
    nuovi_valori = data.get("valori")

    if row_index is None or not isinstance(nuovi_valori, list):
        return jsonify({"error": "Parametri mancanti"}), 400

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "Nessun dato da modificare"}), 404

    try:
        _modifica_excel(excel_path, row_index, nuovi_valori)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@clienti.route("/<id_cliente>/elimina-riga", methods=["POST"])
@login_required
@staff_required
def elimina_riga(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        return jsonify({"error": "Cliente non trovato"}), 404

    data = request.get_json(silent=True) or {}
    row_index = data.get("row_index")

    if row_index is None:
        return jsonify({"error": "Parametri mancanti"}), 400

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "Nessun dato da eliminare"}), 404

    try:
        _elimina_excel(excel_path, row_index)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@clienti.route("/importa-excel")
@login_required
def import_excel():
    """Pagina per selezionare cliente e importare Excel."""
    plugins = client_loader.get_all_plugins()
    return render_template("clienti/importa_excel.html", plugins=plugins)


@clienti.route("/<id_cliente>/importa-excel", methods=["POST"])
@login_required
def esegui_import_excel(id_cliente):
    """Importa un file Excel per il cliente, aggiornando le giacenze."""
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        flash("Cliente non trovato.", "error")
        return redirect(url_for("clienti.elenco"))

    f = request.files.get("excel_file")
    if not f or not f.filename:
        flash("Nessun file selezionato.", "error")
        return redirect(url_for("clienti.import_excel"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("Il file deve essere in formato .xlsx", "error")
        return redirect(url_for("clienti.import_excel"))

    upload_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "excel_import", id_cliente)
    os.makedirs(upload_dir, exist_ok=True)
    path_excel = os.path.join(upload_dir, secure_filename(f.filename))
    f.save(path_excel)

    try:
        from import_export.excel_importer import importa_excel_cliente
        stats = importa_excel_cliente(id_cliente, path_excel)
        flash(f"Importato: {stats.get('giacenze_inserite', 0)} giacenze inserite, {stats.get('giacenze_aggiornate', 0)} aggiornate.", "success")
    except Exception as e:
        flash(f"Errore importazione: {str(e)}", "error")
        current_app.logger.exception(f"Import Excel fallito per {id_cliente}")

    return redirect(url_for("clienti.dettaglio", id_cliente=id_cliente))


@clienti.route("/<id_cliente>/aggiungi-riga", methods=["POST"])
@login_required
def aggiungi_riga(id_cliente):
    plugin = client_loader.get_plugin(id=id_cliente)
    if not plugin:
        return jsonify({"error": "Cliente non trovato"}), 404

    data = request.get_json(silent=True) or {}
    nuovi_valori = data.get("valori")

    if not isinstance(nuovi_valori, list):
        return jsonify({"error": "Parametri mancanti"}), 400

    excel_dir = _get_excel_dir()
    excel_path = os.path.join(excel_dir, f"{id_cliente}.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "Nessun file Excel"}), 404

    try:
        _aggiungi_excel(excel_path, nuovi_valori)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500