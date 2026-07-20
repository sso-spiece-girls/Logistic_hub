from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.excel_writer import CENTER_ALIGN

BBLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GRAY = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
WHITE_F = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NORM_F = Font(name="Calibri", size=10)
BOLD_F = Font(name="Calibri", size=10, bold=True)
THIN_B = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))


def _hdr(ws, row, cols, fill=BBLUE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = WHITE_F
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_B


class EneganExcelWriter:
    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        wb = Workbook()
        self._crea_uscite(wb, ddt_data_list)
        self._crea_ingresso(wb)
        wb.save(excel_path)
        return excel_path

    def _crea_uscite(self, wb, ddt_data_list):
        ws = wb.active
        ws.title = "USCITE"
        ws.cell(row=5, column=1, value="DEPOSITO COLLESALVETTI").font = BOLD_F
        ws.cell(row=5, column=6, value="legenda costo materiali è :").font = BOLD_F
        ws.cell(row=5, column=7, value="DESCRIZIONE").font = BOLD_F
        ws.cell(row=5, column=8, value="TARIFFA").font = BOLD_F
        ws.cell(row=6, column=7, value="€/Master carton")
        ws.cell(row=6, column=8, value=0.5)
        ws.cell(row=7, column=7, value="€/Pz. Singolo")
        ws.cell(row=7, column=8, value=0.25)
        ws.cell(row=8, column=1, value="USCITE MESE :")
        ws.cell(row=8, column=2, value="GIUGNO 2026")
        ws.cell(row=8, column=7, value="€/Minimo tassabile")
        ws.cell(row=8, column=8, value=3)
        ws.cell(row=9, column=7, value="€/Kit omaggio")
        ws.cell(row=9, column=8, value=1)
        ws.cell(row=9, column=11, value="COSTO Coop SQ Totale (70% del ricavo)")
        ws.cell(row=10, column=7, value="Deposito mensile")
        ws.cell(row=10, column=8, value=1000)

        hdr = ["Data uscita merce", "DDT", "VS. CLIENTE", "Master carton",
               "Pz. sfusi", "KIT omaggio", "Importo Totale", "GESTIONE DOCUMENTI",
               "Note", "", "", "calcolo importo Colli", "verifica se Minimo Tassabile"]
        for i, h in enumerate(hdr, 1):
            ws.cell(row=11, column=i, value=h)
        _hdr(ws, 11, 14)
        for col, w in zip("ABCDEFGHIJKLMN", [16, 10, 22, 10, 10, 10, 12, 12, 20, 5, 5, 14, 10]):
            ws.column_dimensions[col].width = w

        data_start = 12
        for i, ddt in enumerate(ddt_data_list):
            r = data_start + i
            ws.cell(row=r, column=1, value=ddt.get("data", "")).alignment = CENTER_ALIGN
            ws.cell(row=r, column=2, value=ddt.get("ddt", "")).alignment = CENTER_ALIGN
            ws.cell(row=r, column=3, value=ddt.get("cliente", ""))
            master = ddt.get("extra", {}).get("master_carton", 0)
            pz = ddt.get("extra", {}).get("pz_sfusi", 0)
            kit = ddt.get("extra", {}).get("kit", 0)
            ws.cell(row=r, column=4, value=master).alignment = CENTER_ALIGN
            ws.cell(row=r, column=5, value=pz).alignment = CENTER_ALIGN
            ws.cell(row=r, column=6, value=kit).alignment = CENTER_ALIGN
            ws.cell(row=r, column=13, value=f"=D{r}*0.5+E{r}*0.25+F{r}*1")
            ws.cell(row=r, column=14, value=f"=IF(M{r}<3,3,M{r})")
            ws.cell(row=r, column=7, value=f"=IF(A{r}>0,N{r},0)")
            ws.cell(row=r, column=8, value=1).alignment = CENTER_ALIGN
            ws.cell(row=r, column=11, value=f"=IF(A{r}>0,N{r}*0.7,0)")

        tot_r = data_start + len(ddt_data_list)
        ws.cell(row=tot_r, column=6, value="TOTALE USCITE").font = BOLD_F
        ws.cell(row=tot_r, column=7, value=f"=SUM(G12:G{tot_r-1})").font = BOLD_F
        ws.cell(row=tot_r, column=8, value=f"=SUM(H12:H{tot_r-1})").font = BOLD_F
        ws.cell(row=tot_r, column=11, value=f"=SUM(K12:K{tot_r-1})").font = BOLD_F

    def _crea_ingresso(self, wb):
        ws = wb.create_sheet("INGRESSO")
        ws.cell(row=1, column=7, value="LEGENDA COSTI").font = BOLD_F
        ws.cell(row=2, column=7, value="DESCRIZIONE").font = BOLD_F
        ws.cell(row=2, column=8, value="COSTO").font = BOLD_F
        ws.cell(row=3, column=7, value="Entrata Master carton")
        ws.cell(row=3, column=8, value=0.5)
        ws.cell(row=4, column=7, value="Entrata a pz. Sfuso")
        ws.cell(row=4, column=8, value=0.2)
        ws.cell(row=5, column=7, value="Minimo tassabile")
        ws.cell(row=5, column=8, value=2.75)
        ws.cell(row=6, column=1, value="DEPOSITO COLLESALVETTI").font = BOLD_F
        ws.cell(row=6, column=7, value="KIT Resi a pacchetto")
        ws.cell(row=6, column=8, value=1)
        ws.cell(row=7, column=7, value="Resi da Vendite  a pz.")
        ws.cell(row=7, column=8, value=0.24)
        ws.cell(row=8, column=1, value="MESE :")
        ws.cell(row=8, column=2, value="GIUGNO 2026")

        hdr = ["DATA", "N. Doc Trasporto", "Fornitore / mittente",
               "Master carton", "Pz. sfusi", "Resi KIT omaggio",
               "Resi da vendita", "IMPORTO TOTALE",
               "", "", "", "calcolo importo Colli", "verifica se Minimo Tassabile"]
        for i, h in enumerate(hdr, 1):
            ws.cell(row=14, column=i, value=h)
        _hdr(ws, 14, 14)
        for col, w in zip("ABCDEFGHIJKLMN", [12, 14, 22, 10, 10, 10, 10, 12, 5, 5, 5, 14, 10]):
            ws.column_dimensions[col].width = w

        for r in range(15, 25):
            ws.cell(row=r, column=12, value=f"=D{r}*0.5+E{r}*0.2+F{r}*1+G{r}*0.24")
            ws.cell(row=r, column=13, value=f"=IF(L{r}<2.75,2.75,L{r})")
            ws.cell(row=r, column=11, value=f"=IF(A{r}>0,N{r}*0.7,0)")
            ws.cell(row=r, column=8, value=f"=N{r}")

        ws.cell(row=25, column=1, value="TOTALI").font = BOLD_F
        for col_l in ["D", "E", "F", "G"]:
            ws.cell(row=25, column=ord(col_l) - 64, value=f"=SUM({col_l}15:{col_l}23)").font = BOLD_F
        ws.cell(row=26, column=1, value="TOTALE COSTI ENTRATE").font = BOLD_F
        ws.cell(row=26, column=8, value=f"=SUM(H15:H25)").font = BOLD_F

        ws.cell(row=37, column=1, value="DEPOSITO MENSILE").font = BOLD_F
        ws.cell(row=39, column=1, value="MESE").font = BOLD_F
        ws.cell(row=39, column=8, value=1000)
        ws.cell(row=40, column=1, value="PALLET")
        ws.cell(row=40, column=2, value="POSTI PALLET € 5 CAD.")
        ws.cell(row=40, column=8, value=1250)
        ws.cell(row=41, column=1, value="TOTALE DEPOSITO").font = BOLD_F
        ws.cell(row=41, column=8, value=f"=SUM(H38:H40)").font = BOLD_F
