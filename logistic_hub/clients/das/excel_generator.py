from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.excel_writer import CENTER_ALIGN

BBLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WHITE_F = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NORM_F = Font(name="Calibri", size=10)
BOLD_F = Font(name="Calibri", size=10, bold=True)
THIN_B = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))


def _hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = BBLUE
        cell.font = WHITE_F
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_B


MESI = [
    "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
    "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE"
]


class DasExcelWriter:
    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        wb = Workbook()
        wb.remove(wb.active)
        for month in ["LUGLIO 2025", "AGOSTO 2025", "SETTEMBRE",
                       "OTTOBRE", "NOVEMBRE", "DICEMBRE",
                       "GENNAIO", "FEBBRAIO", "MARZO",
                       "APRILE 2026", "MAGGIO 2026", " GIUGNO 2026"]:
            self._crea_mese(wb, month)
        wb.save(excel_path)
        return excel_path

    def _crea_mese(self, wb, nome_mese):
        ws = wb.create_sheet(nome_mese)
        ws.cell(row=1, column=1, value="LOGISTICS  SOLUTION").font = BOLD_F
        ws.cell(row=2, column=1, value="Piattaforma: San miniato")
        ws.cell(row=2, column=2, value="LIMITE").font = BOLD_F
        ws.cell(row=3, column=1, value="Spett.le:").font = BOLD_F
        ws.cell(row=3, column=2, value="DASEUROPE")
        ws.cell(row=4, column=1, value=f"Attività svolto nel mese: {nome_mese.upper()}")
        ws.cell(row=5, column=1, value="Presenti dal mese precedente n° plt")
        ws.cell(row=5, column=3, value=0)

        ws.cell(row=6, column=1, value="DATA")
        ws.cell(row=6, column=2, value="ENTRATE")
        ws.cell(row=6, column=4, value="N. SCARICHI")
        ws.cell(row=6, column=5, value="plt entrati")
        ws.cell(row=6, column=6, value="N.COLLI X CONT.")

        sub = ["GIORNO DEL MESE", "N. DDT", "3,00 PLT ENTRATI",
               "", "3,00 PLT USCITI", "COSTO PER PLT USCITI", "PLT", "5,5 PLT X MESE."]
        for i, h in enumerate(sub, 1):
            ws.cell(row=7, column=i, value=h)
        _hdr(ws, 6, 10)
        _hdr(ws, 7, 10)

        for col, w in zip("ABCDEFGHIJ", [8, 12, 12, 8, 12, 12, 14, 14, 14, 14]):
            ws.column_dimensions[col].width = w

        for day in range(1, 32):
            r = day + 7
            ws.cell(row=r, column=1, value=day).alignment = CENTER_ALIGN
            ws.cell(row=r, column=3, value="").alignment = CENTER_ALIGN
            ws.cell(row=r, column=5, value="").alignment = CENTER_ALIGN
            ws.cell(row=r, column=8, value=f"=G{r}*3")
            if day == 1:
                ws.cell(row=r, column=9, value=f"=C5+C{r}")
            else:
                ws.cell(row=r, column=9, value=f"=I{r-1}+C{r}")
            ws.cell(row=r, column=10, value=f"=I{r}*5.5")

        tot_r = 39
        for col_l in ["C", "D", "E", "F", "G", "H", "I"]:
            col_n = ord(col_l) - 64
            ws.cell(row=tot_r, column=col_n, value=f"=SUM({col_l}8:{col_l}38)" if col_l != "E" else f"=SUM(E7:E38)")
            ws.cell(row=tot_r, column=col_n).font = BOLD_F
        ws.cell(row=tot_r, column=10, value=f"=J38").font = BOLD_F

        ws.cell(row=41, column=1, value="RIEPILOGO").font = BOLD_F
        ws.cell(row=41, column=2, value="IMPORTO").font = BOLD_F
        ws.cell(row=41, column=3, value="QUANTITA").font = BOLD_F
        ws.cell(row=42, column=1, value="Deposito")
        ws.cell(row=42, column=2, value=f"=J39")
        ws.cell(row=42, column=8, value="GIACENZA FINE MESE")
        ws.cell(row=42, column=9, value=f"=C5+C39-G39")
        ws.cell(row=43, column=1, value="PALLETS USCITI")
        ws.cell(row=43, column=2, value=f"=C43*3")
        ws.cell(row=43, column=3, value=f"=G39")
        ws.cell(row=44, column=1, value="N. SCARICHI")
        ws.cell(row=44, column=2, value=f"=C44*150")
        ws.cell(row=44, column=3, value=f"=D39")
        ws.cell(row=45, column=1, value="plt entrati")
        ws.cell(row=45, column=2, value=f"=C45*3")
        ws.cell(row=45, column=3, value=f"=E39")
        ws.cell(row=46, column=1, value="ORE EXTRA")
        ws.cell(row=46, column=2, value=f"=C46*23")
        ws.cell(row=48, column=1, value="TOTALE").font = BOLD_F
        ws.cell(row=48, column=2, value=f"=SUM(B42:B47)").font = BOLD_F
        ws.cell(row=48, column=9, value="TOTALE").font = BOLD_F
        ws.cell(row=48, column=10, value=f"=B48").font = BOLD_F

        ws.cell(row=53, column=15, value="DATA").font = BOLD_F
        ws.cell(row=53, column=16, value="TRASPORTI").font = BOLD_F
