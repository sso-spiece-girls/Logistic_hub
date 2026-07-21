from core.excel_writer import apri_o_crea_excel, stile_intestazione, stile_cella, CENTER_ALIGN, crea_backup
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy


class SoffasExcelWriter:
    """Scrive DDT nel file Soffas — inserisce righe prima delle formule SUM."""

    DATA_START_ROW = 18  # Prima riga dati
    SUM_ROW_OFFSET = 37  # Riga con le formule SUM (prima di inserire)

    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        crea_backup(excel_path)
        wb, esistente = apri_o_crea_excel(excel_path)

        if not esistente:
            self._crea_struttura_base(wb)

        ws = wb["DDT SOFFAS"] if "DDT SOFFAS" in wb.sheetnames else wb.active

        # Determina dove inserire: appena prima della riga SUM (dopo i dati esistenti)
        sum_row = self._trova_riga_somma(ws)

        for ddt in ddt_data_list:
            for art in ddt.get("articoli", []):
                # Inserisci una riga prima del SUM
                ws.insert_rows(sum_row, 1)
                # Copia lo stile della riga precedente
                self._copia_stile_riga(ws, sum_row - 1, sum_row)
                # Scrivi i dati
                stile_cella(ws, sum_row, 3, ddt.get("data", ""), align=CENTER_ALIGN)
                stile_cella(ws, sum_row, 5, art.get("qta", 0), align=CENTER_ALIGN)
                sum_row += 1  # La riga SUM si è spostata in giù

        wb.save(excel_path)
        return excel_path

    def _trova_riga_somma(self, ws):
        """Trova la prima riga che contiene una formula SUM nella colonna C."""
        for row in range(self.SUM_ROW_OFFSET, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=3).value
            if cell_val and isinstance(cell_val, str) and cell_val.startswith("=SUM"):
                return row
        return self.SUM_ROW_OFFSET

    def _copia_stile_riga(self, ws, da_riga, a_riga):
        """Copia lo stile da una riga all'altra."""
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=da_riga, column=col)
            dst = ws.cell(row=a_riga, column=col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format

    def _crea_struttura_base(self, wb):
        """Crea la struttura base se il file non esiste."""
        ws = wb.active
        ws.title = "DDT SOFFAS"
        # Headers
        headers = ["Data", "DDT", "Qualita", "Bobina", "Peso (Kg)", "Pallet", "Note"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        stile_intestazione(ws, 1, len(headers))
        for col, w in zip("ABCDEFG", [14, 16, 20, 20, 14, 10, 30]):
            ws.column_dimensions[col].width = w