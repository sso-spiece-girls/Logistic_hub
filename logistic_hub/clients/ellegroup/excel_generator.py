from core.excel_writer import apri_o_crea_excel, stile_intestazione, stile_cella, CENTER_ALIGN, LEFT_ALIGN, crea_backup
from openpyxl.styles import Font


class EllegroupExcelWriter:
    """Scrive DDT nel file Ellegroup — scrive nelle prime righe vuote disponibili.
    Il file ha righe 15-180 pre-formattate con formule =IF(F>0,F*1.17,0).
    """

    DATA_START_ROW = 15
    DATA_END_ROW = 180
    TOTALS_ROW = 181

    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        crea_backup(excel_path)
        wb, esistente = apri_o_crea_excel(excel_path)

        if not esistente:
            ws = wb.active
            ws.title = "DDT ELLE GROUP"
            headers = ["Data", "DDT", "Agente", "Cliente", "Prov.", "Importo", "Note"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=h)
            stile_intestazione(ws, 1, len(headers))
            for col, w in zip("ABCDEFG", [14, 16, 20, 28, 14, 14, 30]):
                ws.column_dimensions[col].width = w

        ws = wb["DDT ELLE GROUP"] if "DDT ELLE GROUP" in wb.sheetnames else wb.active

        for ddt in ddt_data_list:
            for art in ddt.get("articoli", []):
                # Trova la prima riga vuota nella sezione dati (colonna D = DDT)
                target_row = self._trova_riga_vuota(ws)
                if target_row is None:
                    continue  # Sezione dati piena

                stile_cella(ws, target_row, 1, ddt.get("data", ""), align=CENTER_ALIGN)
                stile_cella(ws, target_row, 2, ddt.get("cliente", ""), align=LEFT_ALIGN)
                stile_cella(ws, target_row, 3, "", align=CENTER_ALIGN)
                stile_cella(ws, target_row, 4, ddt.get("ddt", ""), align=CENTER_ALIGN)
                stile_cella(ws, target_row, 5, art.get("descrizione", ""), align=LEFT_ALIGN)
                stile_cella(ws, target_row, 6, art.get("qta", 0), align=CENTER_ALIGN)
                # Col G ha già la formula =IF(F>0,F*1.17,0) — non toccarla
                stile_cella(ws, target_row, 8, art.get("qta", 0), align=CENTER_ALIGN)

        wb.save(excel_path)
        return excel_path

    def _trova_riga_vuota(self, ws):
        """Trova la prima riga vuota nella sezione dati (col D senza valore)."""
        for row in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
            d_val = ws.cell(row=row, column=4).value  # Col D = DDT
            if d_val is None or str(d_val).strip() == "":
                return row
        return None  # Sezione piena