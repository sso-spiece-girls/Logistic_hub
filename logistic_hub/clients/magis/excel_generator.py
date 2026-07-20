from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.excel_writer import stile_intestazione, CENTER_ALIGN

BBLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WHITE_F = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NORM_F = Font(name="Calibri", size=10)
BOLD_F = Font(name="Calibri", size=10, bold=True)
THIN_B = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = BBLUE
        cell.font = WHITE_F
        cell.alignment = CENTER
        cell.border = THIN_B


class MagisExcelWriter:
    def __init__(self, config):
        self.cfg = config

    def genera_excel(self, ddt_data_list, excel_path):
        wb = Workbook()
        self._crea_big_bag(wb, ddt_data_list)
        self._crea_small_bag(wb, ddt_data_list)
        self._crea_giacenze(wb)
        wb.save(excel_path)
        return excel_path

    def _crea_big_bag(self, wb, ddt_data_list):
        ws = wb.active
        ws.title = "MARZO - BIG BAG"
        ws.cell(row=1, column=1, value="LOGISTIC  SOLUTION")
        ws.cell(row=2, column=1, value="Spett.le:").font = BOLD_F
        ws.cell(row=2, column=2, value="MAGIS")
        ws.cell(row=3, column=1, value="Attività svolto nel mese: MARZO 2026")
        ws.cell(row=4, column=1, value="Presenti dal mese precedente n° plt")
        ws.cell(row=4, column=3, value=728)
        headers = ["DATA", "ENTRATE", "", "GIACENZE PANCALE  A"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=5, column=i, value=h)
        sub = ["GIORNO DEL MESE", "N. container", "PLT IN", "PLT OUT",
               "COSTO PER SCARICO PALLETS", "COSTO PER PLT USCITI", "PLT", "0,117TOT. PLT X GG."]
        for i, h in enumerate(sub, 1):
            ws.cell(row=6, column=i, value=h)
        _style_hdr(ws, 5, 8)
        _style_hdr(ws, 6, 8)
        for c, w in zip("ABCDEFGH", [8, 28, 8, 8, 14, 14, 8, 14]):
            ws.column_dimensions[c].width = w

        for day in range(1, 32):
            r = day + 6
            ws.cell(row=r, column=1, value=day).alignment = CENTER
            ws.cell(row=r, column=2, value="").alignment = CENTER
            ws.cell(row=r, column=3, value="").alignment = CENTER
            ws.cell(row=r, column=4, value="").alignment = CENTER
            ws.cell(row=r, column=5, value=f"=C{r}*2.815")
            ws.cell(row=r, column=6, value=f"=D{r}*2.815")
            if day == 1:
                ws.cell(row=r, column=7, value=f"=C4+C{r}-D{r}")
            else:
                ws.cell(row=r, column=7, value=f"=G{r-1}+C{r}-D{r}")
            ws.cell(row=r, column=8, value=f"=G{r}*0.117")

        tot_r = 38
        for col_letter in ["C", "D", "E", "F", "G", "H"]:
            ws.cell(row=tot_r, column=ord(col_letter) - 64,
                    value=f"=SUM({col_letter}7:{col_letter}37)")
            ws.cell(row=tot_r, column=ord(col_letter) - 64).font = BOLD_F

        ws.cell(row=40, column=1, value="RIEPILOGO").font = BOLD_F
        ws.cell(row=40, column=2, value="IMPORTO").font = BOLD_F
        ws.cell(row=40, column=3, value="QUANTITA").font = BOLD_F
        ws.cell(row=41, column=1, value="PALLETS ENTRATI")
        ws.cell(row=41, column=2, value=f"=C41*2.815")
        ws.cell(row=41, column=3, value=f"=C38")
        ws.cell(row=42, column=1, value="Deposito")
        ws.cell(row=42, column=2, value=f"=0.117*G38")
        ws.cell(row=42, column=3, value=f"=G38")
        ws.cell(row=43, column=1, value="PALLETS usciti")
        ws.cell(row=43, column=2, value=f"=C43*2.815")
        ws.cell(row=43, column=3, value=f"=D38")
        ws.cell(row=44, column=1, value="TOTALE").font = BOLD_F
        ws.cell(row=44, column=2, value=f"=B41+B42+B43").font = BOLD_F
        ws.cell(row=44, column=7, value="TOTALE").font = BOLD_F
        ws.cell(row=44, column=8, value=f"=B44").font = BOLD_F

    def _crea_small_bag(self, wb, ddt_data_list):
        ws = wb.create_sheet("MARZO-SMALL BAG")
        ws.cell(row=1, column=1, value="LOGISTICS  SOLUTION")
        ws.cell(row=2, column=1, value="Spett.le:").font = BOLD_F
        ws.cell(row=2, column=2, value="MAGIS")
        ws.cell(row=3, column=1, value="Attività svolto nel mese: MARZO2026")
        ws.cell(row=4, column=1, value="Presenti dal mese precedente n° plt")
        ws.cell(row=4, column=3, value=189)
        headers = ["DATA", "ENTRATE", "", "GIACENZE PANCALE  B"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=5, column=i, value=h)
        sub = ["GIORNO DEL MESE", "N. container", "PLT IN", "PLT OUT",
               "COSTO PER SCARICO PALLETS", "COSTO PER PLT USCITI", "PLT", "0,253 TOT. PLT X GG."]
        for i, h in enumerate(sub, 1):
            ws.cell(row=6, column=i, value=h)
        _style_hdr(ws, 5, 8)
        _style_hdr(ws, 6, 8)
        for c, w in zip("ABCDEFGH", [8, 28, 8, 8, 14, 14, 8, 14]):
            ws.column_dimensions[c].width = w

        for day in range(1, 32):
            r = day + 6
            ws.cell(row=r, column=1, value=day).alignment = CENTER
            ws.cell(row=r, column=2, value="").alignment = CENTER
            ws.cell(row=r, column=3, value="").alignment = CENTER
            ws.cell(row=r, column=4, value="").alignment = CENTER
            ws.cell(row=r, column=5, value=f"=C{r}*2.815")
            ws.cell(row=r, column=6, value=f"=D{r}*2.815")
            if day == 1:
                ws.cell(row=r, column=7, value=f"=ROUND(C4+C{r}-D{r},0)")
            else:
                ws.cell(row=r, column=7, value=f"=G{r-1}+C{r}-D{r}")
            ws.cell(row=r, column=8, value=f"=G{r}*0.253")

        tot_r = 38
        for col_letter in ["C", "D", "E", "F", "G", "H"]:
            ws.cell(row=tot_r, column=ord(col_letter) - 64,
                    value=f"=SUM({col_letter}7:{col_letter}37)")
            ws.cell(row=tot_r, column=ord(col_letter) - 64).font = BOLD_F

        ws.cell(row=40, column=1, value="RIEPILOGO").font = BOLD_F
        ws.cell(row=40, column=2, value="IMPORTO").font = BOLD_F
        ws.cell(row=40, column=3, value="QUANTITA").font = BOLD_F
        ws.cell(row=41, column=1, value="PALLETS ENTRATI")
        ws.cell(row=41, column=2, value=f"=C41*2.815")
        ws.cell(row=41, column=3, value=f"=C38")
        ws.cell(row=42, column=1, value="Deposito")
        ws.cell(row=42, column=2, value=f"=0.253*G38")
        ws.cell(row=42, column=3, value=f"=G38")
        ws.cell(row=43, column=1, value="Plts usciti")
        ws.cell(row=43, column=2, value=f"=C43*2.815")
        ws.cell(row=43, column=3, value=f"=D38")
        ws.cell(row=47, column=1, value="TOTALE").font = BOLD_F
        ws.cell(row=47, column=2, value=f"=SUM(B41:B46)").font = BOLD_F
        ws.cell(row=47, column=7, value="TOTALE").font = BOLD_F
        ws.cell(row=47, column=8, value=f"=SUM(B41:B46)").font = BOLD_F

    def _crea_giacenze(self, wb):
        ws = wb.create_sheet("GIACENZE MARZO")
        ws.cell(row=2, column=4, value="DAL 13/06/2023")
        h1 = ["Codice", "Confezione", "Materiale", "DEP. LIMITE", "USCITE", "SALDO",
              "", "Codice", "Confezione", "Materiale", "DEP.SOVIGLIANA", "USCITE", "SALDO"]
        for i, h in enumerate(h1, 1):
            ws.cell(row=4, column=i, value=h)

        data = [
            ("JH8152", "big bag", "gomma", 3315, 3315, "", "", "", "", ""),
            ("JH8152", "small bag", "gomma", 1195, 1195, "", "", "", "", ""),
            ("JH3200", "big bag", "resina", 2836, 2836, "", "", "", "", ""),
            ("JH3200", "small bag", "resina", 509, 509, "", "", "", "", ""),
            ("JH3204", "small bag", "resina", 302, 302, "", "", "", "", ""),
            ("JH3204", "big bag", "resina", 452, 452, "", "", "", "", ""),
        ]
        for i, (cod, conf, mat, dep, usc, _, _, _, _, _) in enumerate(data):
            r = 5 + i
            ws.cell(row=r, column=1, value=cod).alignment = CENTER
            ws.cell(row=r, column=2, value=conf).alignment = CENTER
            ws.cell(row=r, column=3, value=mat).alignment = CENTER
            ws.cell(row=r, column=4, value=dep).alignment = CENTER
            ws.cell(row=r, column=5, value=usc).alignment = CENTER
            ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        ws.cell(row=11, column=4, value="=SUM(D5:D10)")
        ws.cell(row=11, column=5, value="=SUM(E5:E10)")
        ws.cell(row=11, column=6, value="=SUM(F5:F10)")

        h2 = ["Codice", "Confezione", "Materiale", "DEP. Montelupo", "USCITE", "SALDO",
              "", "Codice", "Confezione", "Materiale", "DEP. CAMBIANO", "USCITE", "SALDO"]
        for i, h in enumerate(h2, 1):
            ws.cell(row=14, column=i, value=h)

        data2 = [
            ("JH8152", "big bag", "gomma", 0, 0, "", "JH8152", "big bag", "gomma", 2595, 2050),
            ("JH8152", "small bag", "gomma", 0, 0, "", "JH8152", "small bag", "gomma", 676, 521),
            ("JH3200", "big bag", "resina", 0, 0, "", "JH3200", "big bag", "resina", 2065, 1547),
            ("JH3200", "small bag", "resina", 0, 0, "", "JH3200", "small bag", "resina", 290, 177),
            ("JH3204", "small bag", "resina", 0, 0, "", "JH3204", "small bag", "resina", 383, 306),
        ]
        for i, row_data in enumerate(data2):
            r = 15 + i
            for j, v in enumerate(row_data):
                if v:
                    ws.cell(row=r, column=j + 1, value=v).alignment = CENTER
            ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
            ws.cell(row=r, column=13, value=f"=K{r}-L{r}")
