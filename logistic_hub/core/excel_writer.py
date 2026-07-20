import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BLUE_FILL = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def stile_intestazione(ws, riga, max_col):
    """Applica stile header blu a una riga."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=riga, column=col)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def stile_cella(ws, riga, col, valore, bold=False, align=None):
    """Scrive una cella con stile."""
    cell = ws.cell(row=riga, column=col, value=valore)
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.alignment = align or CENTER_ALIGN
    cell.border = THIN_BORDER
    if riga % 2 == 0:
        cell.fill = LIGHT_FILL
    return cell


def apri_o_crea_excel(percorso):
    """Apre Excel esistente o ne crea uno nuovo."""
    if os.path.exists(percorso):
        wb = load_workbook(percorso)
        return wb, True
    wb = Workbook()
    return wb, False


def crea_backup(percorso):
    """Crea backup .backup del file Excel prima di modificarlo."""
    if os.path.exists(percorso):
        backup = percorso + ".backup"
        import shutil
        shutil.copy2(percorso, backup)
        return backup
    return None