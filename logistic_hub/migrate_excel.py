"""
Script di migrazione dati dal file Excel GIACENZABOBINECELTEXCOLLE.xlsx
nel database logistic_hub.db.

Utilizzo:
    python migrate_excel.py --excel "path/to/file.xlsx" [--db "path/to/database.db"]

Il foglio GIACENZA deve avere colonne: COD. ARTICOLO, ID BOBINA, PESO,
DESCRIZIONE, QUALITÀ, STATO, UBICAZIONE, PROVENIENZA
"""

import os
import sys
import argparse

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask
from extensions import db
from models import Giacenza, User
from config import Config


def create_app(db_path=None):
    app = Flask(__name__)
    app.config.from_object(Config)
    if db_path:
        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
    db.init_app(app)
    return app


def migrate(excel_path, db_path=None, sheet_name="GIACENZA", force=False):
    try:
        import openpyxl
    except ImportError:
        print("[ERRORE] openpyxl non installato. Esegui: pip install openpyxl")
        return

    if not os.path.exists(excel_path):
        print(f"[ERRORE] File non trovato: {excel_path}")
        return

    app = create_app(db_path)

    with app.app_context():
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Fogli disponibili: {wb.sheetnames}")
            print(f"[ERRORE] Foglio '{sheet_name}' non trovato.")
            return

        ws = wb[sheet_name]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Mappa nomi colonne (case-insensitive)
        col_map = {}
        for i, h in enumerate(header):
            if h:
                key = h.strip().upper().replace(" ", "_")
                if "COD" in key and "ARTICOLO" in key:
                    col_map["codice_articolo"] = i
                elif "ID" in key and "BOBINA" in key:
                    col_map["id_bobina"] = i
                elif key == "PESO":
                    col_map["peso_kg"] = i
                elif "DESCRIZIONE" in key:
                    col_map["descrizione"] = i
                elif "QUALIT" in key:
                    col_map["qualita"] = i
                elif key == "STATO":
                    col_map["stato"] = i
                elif "UBICAZIONE" in key:
                    col_map["ubicazione"] = i
                elif "PROVENIENZA" in key:
                    col_map["provenienza"] = i

        print(f"Colonne riconosciute: {col_map}")
        if "codice_articolo" not in col_map:
            print(f"[ERRORE] Colonna 'COD. ARTICOLO' non trovata. Header: {header}")
            return

        user = User.query.filter_by(username="Francesco").first()
        user_id = user.id if user else None

        importate = 0
        saltate = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            codice = str(row[col_map["codice_articolo"]]).strip() if row[col_map["codice_articolo"]] else ""
            if not codice or codice == "None":
                saltate += 1
                continue

            id_bobina = str(row[col_map.get("id_bobina", -1)]).strip() if col_map.get("id_bobina", -1) >= 0 and row[col_map["id_bobina"]] else None
            if id_bobina == "None":
                id_bobina = None

            if id_bobina:
                esiste = Giacenza.query.filter_by(id_bobina=id_bobina).first()
                if esiste and not force:
                    saltate += 1
                    continue

            peso_str = str(row[col_map.get("peso_kg", -1)] or "0").replace(",", ".")
            try:
                peso = float(peso_str) if peso_str != "None" else 0
            except ValueError:
                peso = 0

            giacenza = Giacenza(
                codice_articolo=codice,
                descrizione=str(row[col_map.get("descrizione", -1)] or codice) if col_map.get("descrizione", -1) >= 0 else codice,
                quantita=0,
                colli=1,
                pallet=0,
                peso_kg=peso,
                id_bobina=id_bobina,
                qualita=str(row[col_map.get("qualita", -1)] or "") if col_map.get("qualita", -1) >= 0 else "",
                provenienza=str(row[col_map.get("provenienza", -1)] or "") if col_map.get("provenienza", -1) >= 0 else "",
                ubicazione=str(row[col_map.get("ubicazione", -1)] or "") if col_map.get("ubicazione", -1) >= 0 else "",
                magazzino="Colle 4",
                updated_by=user_id,
            )
            db.session.add(giacenza)
            importate += 1

        db.session.commit()
        print(f"Migrazione completata: {importate} bobine importate, {saltate} saltate (già esistenti o vuote)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migra dati Excel giacenze bobine")
    parser.add_argument("--excel", required=True, help="Percorso file Excel")
    parser.add_argument("--db", help="Percorso database SQLite (default: da config)")
    parser.add_argument("--sheet", default="GIACENZA", help="Nome foglio Excel (default: GIACENZA)")
    parser.add_argument("--force", action="store_true", help="Forza aggiornamento anche se già presenti")
    args = parser.parse_args()

    migrate(args.excel, args.db, args.sheet, args.force)