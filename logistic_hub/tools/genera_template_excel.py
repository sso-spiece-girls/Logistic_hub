"""
Genera i file template Excel per tutti i clienti nella cartella excel_clienti.
I template servono come base strutturata per l'append dei DDT durante l'upload.
"""
import sys, os

# Add project to path
PROJECT = r"C:\Users\marzu\OneDrive\Desktop\Programmi_Per_Logistic\LogisticProgrammaTotale\logistic_hub"
sys.path.insert(0, PROJECT)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Style constants (same as core/excel_writer.py)
BLUE_FILL = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

EXCEL_DIR = os.path.join(PROJECT, "data", "docs", "excel_clienti")
os.makedirs(EXCEL_DIR, exist_ok=True)


def stile_intestazione(ws, riga, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=riga, column=col)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def crea_template(filename, sheet_name, headers, column_widths):
    path = os.path.join(EXCEL_DIR, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    stile_intestazione(ws, 1, len(headers))

    for letter, w in column_widths:
        ws.column_dimensions[letter].width = w

    wb.save(path)
    print(f"  [OK] {filename} (sheet: '{sheet_name}', cols: {len(headers)})")
    return path


def main():
    print("Generazione template Excel per tutti i clienti...\n")

    # === DAS EUROPE ===
    crea_template("das.xlsx", "DDT DAS",
        ["Data", "DDT", "Pallet ID", "Cliente", "Indirizzo", "Peso (Kg)", "Note"],
        [("A", 14), ("B", 22), ("C", 20), ("D", 25), ("E", 30), ("F", 12), ("G", 30)])

    # === ELLE GROUP ===
    crea_template("ellegroup.xlsx", "DDT ELLE GROUP",
        ["Data", "DDT", "Agente", "Cliente", "Prov.", "Importo", "Note"],
        [("A", 14), ("B", 22), ("C", 20), ("D", 28), ("E", 8), ("F", 14), ("G", 30)])

    # === ENEGAN SPA ===
    crea_template("enegan.xlsx", "DDT ENEGAN",
        ["Data", "DDT", "Cliente DDT", "Cliente Finale", "Articolo", "Qta", "Unità", "Note"],
        [("A", 14), ("B", 22), ("C", 25), ("D", 28), ("E", 16), ("F", 10), ("G", 8), ("H", 30)])

    # === LA LECCIA ===
    crea_template("laleccia.xlsx", "DDT LA LECCIA",
        ["Data", "DDT", "Ristorante", "Articolo", "Qta", "Importo", "Note"],
        [("A", 14), ("B", 16), ("C", 28), ("D", 30), ("E", 10), ("F", 14), ("G", 30)])

    # === MAGIS SPA ===
    magis_path = os.path.join(EXCEL_DIR, "magis.xlsx")
    wb = Workbook()
    for i, (name, headers, widths) in enumerate([
        ("BIG BAG", ["Data", "DDT", "Cliente", "PLT OUT", "Note"],
         [("A", 14), ("B", 20), ("C", 25), ("D", 12), ("E", 30)]),
        ("SMALL BAG", ["Data", "DDT", "Cliente", "PLT OUT", "Note"],
         [("A", 14), ("B", 20), ("C", 25), ("D", 12), ("E", 30)]),
        ("GIACENZE", ["Deposito", "Articolo", "Qta Iniziale", "Uscite", "Qta Finale"],
         [("A", 14), ("B", 25), ("C", 16), ("D", 12), ("E", 16)]),
    ]):
        if i == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = name
        for j, h in enumerate(headers, 1):
            ws.cell(row=1, column=j, value=h)
        stile_intestazione(ws, 1, len(headers))
        for letter, w in widths:
            ws.column_dimensions[letter].width = w
    wb.save(magis_path)
    print(f"  [OK] magis.xlsx (sheets: 'BIG BAG', 'SMALL BAG', 'GIACENZE')")

    # === SOFFAS ===
    crea_template("soffas.xlsx", "DDT SOFFAS",
        ["Data", "DDT", "Qualità", "Bobina", "Peso (Kg)", "Pallet", "Note"],
        [("A", 14), ("B", 22), ("C", 20), ("D", 18), ("E", 12), ("F", 10), ("G", 30)])

    print(f"\nTutti i template creati in: {EXCEL_DIR}")
    print("\nFile generati:")
    for f in sorted(os.listdir(EXCEL_DIR)):
        if f.endswith(".xlsx"):
            path = os.path.join(EXCEL_DIR, f)
            wb_tmp = __import__('openpyxl').load_workbook(path, read_only=True)
            sheets = ", ".join(wb_tmp.sheetnames)
            wb_tmp.close()
            print(f"  {f:30s} ({sheets})")


if __name__ == "__main__":
    main()
