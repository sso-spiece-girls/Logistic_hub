import os, re
from datetime import datetime
from openpyxl import load_workbook
from extensions import db
from models import Giacenza


def importa_excel_cliente(id_cliente, excel_path):
    importers = {
        "soffas": _import_soffas,
        "magis": _import_magis,
        "das": _import_das,
        "ellegroup": _import_ellegroup,
        "enegan": _import_enegan,
        "laleccia": _import_laleccia,
    }
    fn = importers.get(id_cliente)
    if not fn:
        raise ValueError(f"Nessun import registrato per il cliente '{id_cliente}'")
    stats = fn(excel_path)
    # Fallback: se nessuna giacenza importata, prova formato riepilogo
    if id_cliente == "soffas" and stats.get("giacenze_inserite", 0) == 0 and stats.get("giacenze_aggiornate", 0) == 0:
        stats2 = _import_soffas_riepilogo(excel_path)
        stats["giacenze_inserite"] += stats2.get("giacenze_inserite", 0)
        stats["giacenze_aggiornate"] += stats2.get("giacenze_aggiornate", 0)
    return stats


def _upsert(codice, descrizione, quantita=0, colli=0, pallet=0, stats=None):
    if stats is None:
        stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}
    existing = Giacenza.query.filter_by(codice_articolo=codice).first()
    if existing:
        existing.quantita = quantita
        existing.colli = colli
        existing.pallet = pallet
        stats["giacenze_aggiornate"] += 1
    else:
        g = Giacenza(
            codice_articolo=codice, descrizione=descrizione,
            quantita=quantita, colli=colli, pallet=pallet,
        )
        db.session.add(g)
        stats["giacenze_inserite"] += 1
    return stats


def _n(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return 0


# ─── SOFFAS ───────────────────────────────────────────────────────────────
# File: BOBINE SOFFASS - LIMITE (1).xlsx
# R3 col 4 (D3) di ogni foglio mese = pallet in deposito
# R33 col 3 = totale entrati mese, R33 col 5 = totale usciti mese
# L'ultimo foglio (LUGLIO) D3 = giacenza attuale

def _import_soffas(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}

    for sname in wb.sheetnames:
        if sname == "Foglio1":
            continue
        ws = wb[sname]
        nome = sname.strip().upper()

        # D3 = pallet in deposito
        giac = _n(ws.cell(row=3, column=4).value)
        # R33: C33 = tot entrati, E33 = tot usciti
        tot_in = _n(ws.cell(row=33, column=3).value) if ws.max_row >= 33 else 0
        tot_out = _n(ws.cell(row=33, column=5).value) if ws.max_row >= 33 else 0

        if giac:
            _upsert(f"SOFFAS_{nome}_DEPOSITO", f"SOFFAS {nome} pallet deposito",
                    quantita=giac, pallet=int(giac), stats=stats)
        if tot_in:
            _upsert(f"SOFFAS_{nome}_ENTRATI", f"SOFFAS {nome} entrati",
                    quantita=tot_in, pallet=int(tot_in), stats=stats)
        if tot_out:
            _upsert(f"SOFFAS_{nome}_USCITI", f"SOFFAS {nome} usciti",
                    quantita=tot_out, pallet=int(tot_out), stats=stats)

    db.session.commit()
    wb.close()
    return stats


# ─── MAGIS ───────────────────────────────────────────────────────────────
# File: FILE MENSILE MAGIS - MARZO 2026.xlsx
# GIACENZE FEBBRAIO / GIACENZE MARZO:
#   Ogni sezione ha un TOTALE in fondo:
#   R31 col 11 = TOTALE (es. 1605 per SPICCHIO in GIACENZE MARZO)
#   R22 col 11 = DEP. CAMBIANO total
#   R11 col 6 = DEP. LIMITE total
# MARZO - BIG BAG / MARZO-SMALL BAG:
#   RIEPILOGO: R41 = PALLETS ENTRATI, R42 = Deposito, R43 = PALLETS usciti

def _import_magis(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}

    for sname in wb.sheetnames:
        ws = wb[sname]

        if sname.startswith("GIACENZE"):
            _leggi_giacenze_magis(ws, sname, stats)
        elif sname in ("MARZO - BIG BAG", "MARZO-SMALL BAG"):
            _leggi_riepilogo_magis(ws, sname, stats)

    db.session.commit()
    wb.close()
    return stats


def _leggi_giacenze_magis(ws, sname, stats):
    sname_clean = sname.replace(" ", "_")
    # Blocco 1: DEP. LIMITE (righe 5-11) + DEP. SOVIGLIANA (colonne 8-11)
    _legge_blocco(ws, 4, 11, "LIMITE", "SOVIGLIANA", sname_clean, stats)
    # Blocco 2: DEP. Montelupo (righe 14-22) + DEP. CAMBIANO
    _legge_blocco(ws, 14, 22, "MONTELUPO", "CAMBIANO", sname_clean, stats)
    # Blocco 3: DEP. SPICCHIO (righe 24-31)
    _legge_blocco(ws, 24, 31, "SPICCHIO", None, sname_clean, stats)


def _legge_blocco(ws, start_row, end_row, depot_sx, depot_dx, sname, stats):
    for r in range(start_row + 1, end_row):
        # Colonna sinistra: codice in A, saldo in F
        cod = ws.cell(row=r, column=1).value
        saldo = ws.cell(row=r, column=6).value
        if cod and isinstance(cod, str) and cod.strip().lower() not in ("codice", "cod."):
            if isinstance(saldo, (int, float)):
                conf = ws.cell(row=r, column=2).value or ""
                mat = ws.cell(row=r, column=3).value or ""
                art = f"{cod}_{conf}" if conf else cod
                _upsert(f"{depot_sx}_{art}", f"{mat} - {conf}" if conf else str(mat or cod),
                        quantita=float(saldo), stats=stats)

        # Colonna destra: codice in H, saldo in K
        if depot_dx:
            cod_dx = ws.cell(row=r, column=8).value
            saldo_dx = ws.cell(row=r, column=11).value
            if cod_dx and isinstance(cod_dx, str) and cod_dx.strip().lower() not in ("codice", "cod."):
                if isinstance(saldo_dx, (int, float)):
                    conf_dx = ws.cell(row=r, column=9).value or ""
                    mat_dx = ws.cell(row=r, column=10).value or ""
                    art_dx = f"{cod_dx}_{conf_dx}" if conf_dx else cod_dx
                    _upsert(f"{depot_dx}_{art_dx}", f"{mat_dx} - {conf_dx}" if conf_dx else str(mat_dx or cod_dx),
                            quantita=float(saldo_dx), stats=stats)

    # TOTALE della sezione
    tot = ws.cell(row=end_row, column=11).value
    if isinstance(tot, (int, float)) and tot:
        _upsert(f"{sname}_{depot_sx}_TOTALE", f"{sname} {depot_sx} totale",
                quantita=float(tot), stats=stats)


def _leggi_riepilogo_magis(ws, sname, stats):
    nome = sname.replace(" ", "_")
    # RIEPILOGO rows (R40-R44)
    riep = {}
    for r in range(40, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        imp = ws.cell(row=r, column=2).value
        qta = ws.cell(row=r, column=3).value
        if label and isinstance(label, str):
            riep[label.strip().upper()] = {"importo": _n(imp), "qta": _n(qta)}

    for key, typ in [("PALLETS ENTRATI", "ENTRATI"), ("Deposito", "DEPOSITO"),
                     ("PALLETS USCITI", "USCITI"), ("TOTALE", "TOTALE")]:
        if key.upper() in riep or key in riep:
            v = riep.get(key.upper()) or riep.get(key)
            if v and v["qta"]:
                _upsert(f"MAGIS_{nome}_{typ}", f"MAGIS {nome} {typ}",
                        quantita=v["qta"], stats=stats)


# ─── DAS ─────────────────────────────────────────────────────────────────
# File: GIACENZE. DAS.xlsx → unico foglio Foglio1: GIAC. in R1609C3
# File: FILE DEPOSITO - DAS FINO A DICEMBRE 2025.xlsx → fogli mensili

def _import_das(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}
    snames = wb.sheetnames

    # GIACENZE. DAS.xlsx → singolo foglio Foglio1
    if snames == ["Foglio1"]:
        ws = wb["Foglio1"]
        h2 = ws.cell(row=2, column=2).value
        if h2 and "Magazzino" in str(h2):
            _import_das_giacenza(ws, stats)
            db.session.commit()
            wb.close()
            return stats


def _import_soffas_riepilogo(excel_path):
    """Formato riepilogo: foglio 'Giacenze' con colonne
    [Codice Articolo, Descrizione, Colli Totali, Peso Totale, Pallet Totali, ...]"""
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        codice = ws.cell(row=r, column=1).value
        desc = ws.cell(row=r, column=2).value
        pallet = _n(ws.cell(row=r, column=5).value)
        if codice and desc:
            _upsert(f"SOFFAS_{codice}", f"SOFFAS {desc}",
                    quantita=pallet, pallet=int(pallet), stats=stats)
    db.session.commit()
    wb.close()
    return stats

    # FILE DEPOSITO → fogli mensili
    for sname in snames:
        ws = wb[sname]
        _import_das_mensile(ws, sname.strip().upper(), stats)

    db.session.commit()
    wb.close()
    return stats


def _import_das_giacenza(ws, stats):
    """Calcola la giacenza dal log pallet: conta entrate - uscite."""
    totale_entrate = 0
    totale_uscite = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        if isinstance(row[0], (int, float)):
            totale_entrate += 1
            if row[8] is not None or row[9] is not None:
                totale_uscite += 1
    giacenza = totale_entrate - totale_uscite
    if giacenza:
        _upsert("DAS_GIACENZA", "DAS giacenza totale",
                quantita=giacenza, pallet=giacenza, stats=stats)
    if totale_entrate:
        _upsert("DAS_TOTALE_ENTRATE", "DAS totale entrate",
                quantita=totale_entrate, pallet=totale_entrate, stats=stats)
    if totale_uscite:
        _upsert("DAS_TOTALE_USCITE", "DAS totale uscite",
                quantita=totale_uscite, pallet=totale_uscite, stats=stats)


def _import_das_mensile(ws, nome_mese, stats):
    presenti_prec = _n(ws.cell(row=5, column=3).value)
    if presenti_prec:
        _upsert(f"DAS_{nome_mese}_PRECEDENTI", f"DAS {nome_mese} presenti prec.",
                quantita=presenti_prec, pallet=int(presenti_prec), stats=stats)

    last = None
    for r in range(ws.max_row, 7, -1):
        if ws.cell(row=r, column=3).value is not None and ws.cell(row=r, column=8).value is not None:
            last = r
            break
    if last:
        for label, col, key in [("ENTRATI", 3, "ENTRATI"), ("USCITI", 6, "USCITI"), ("GIACENZA", 8, "GIACENZA")]:
            v = _n(ws.cell(row=last, column=col).value)
            if v:
                _upsert(f"DAS_{nome_mese}_{key}", f"DAS {nome_mese} {label}",
                        quantita=v, pallet=int(v), stats=stats)


# ─── ELLEGROUP ───────────────────────────────────────────────────────────
# (come prima - funziona gia')

def _import_ellegroup(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}

    if "PROSPETTO USCITE" in wb.sheetnames:
        ws = wb["PROSPETTO USCITE"]
        for row in ws.iter_rows(min_row=15, max_row=ws.max_row, values_only=True):
            ddt_num = row[3]
            if not ddt_num:
                continue
            for agente, cc, cp in [("MARINA", 7, 9), ("GMV", 15, 17), ("MARTA", 23, 25)]:
                colli = row[cc]
                pezzi = row[cp]
                if colli and isinstance(colli, (int, float)):
                    _upsert(f"ELLEGROUP_{ddt_num}_{agente}",
                            f"ELLEGROUP DDT {ddt_num} {agente}",
                            quantita=float(pezzi or 0), colli=int(colli), stats=stats)

    for nome_ingressi in (" PROSPETTO INGRESSI", "PROSPETTO INGRESSI"):
        if nome_ingressi in wb.sheetnames:
            ws = wb[nome_ingressi]
            for row in ws.iter_rows(min_row=15, max_row=ws.max_row, values_only=True):
                cntr = row[0]
                if not cntr:
                    continue
                for agente, cc in [("MARINA_IN", 2), ("GMV_IN", 5), ("MARTA_IN", 8)]:
                    colli = row[cc]
                    if colli and isinstance(colli, (int, float)):
                        _upsert(f"ELLEGROUP_{cntr}_{agente}",
                                f"ELLEGROUP IN {cntr} {agente}",
                                colli=int(colli), stats=stats)
            break

    db.session.commit()
    wb.close()
    return stats


# ─── ENEGAN ──────────────────────────────────────────────────────────────

def _import_enegan(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}
    for sheet_name, tipo in [("USCITE", "uscita"), ("INGRESSO", "ingresso")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True):
            ddt_num = row[1]
            if not ddt_num or not isinstance(ddt_num, (int, float)):
                continue
            mc = _n(row[3]) if len(row) > 3 else 0
            pz = _n(row[4]) if len(row) > 4 else 0
            totale = mc + pz
            if totale:
                _upsert(f"ENEGAN_{ddt_num}_{tipo}",
                        f"ENEGAN DDT {ddt_num} {tipo}",
                        quantita=totale, colli=int(mc), stats=stats)
    db.session.commit()
    wb.close()
    return stats


# ─── LA LECCIA (placeholder) ────────────────────────────────────────────

def _import_laleccia(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    stats = {"giacenze_inserite": 0, "giacenze_aggiornate": 0}
    wb.close()
    return stats
